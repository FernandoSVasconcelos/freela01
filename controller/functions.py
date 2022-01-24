from openpyxl import * 
from numpy import mod, ceil
import numpy as np
from math import *
from tkinter import messagebox

Descod = ['θt', 'U', 'V', 'θf']
aux = 0
nno = 0
Matno = 0  # Dados dos nós
Matbar = 0  # Dados das barras
Carbar = 0  # Dados dos Carregamentos nas Barras
nbar = 0
desl = 0
delta = 0

def Sub_said(matriz, tipo, Nome_aquivo,bar):
    global aux, Descod

    arqexcel = load_workbook(Nome_aquivo)

    # Grava as matrizes locais no arquivo excel
    if tipo == 1:
        Pastatraba = arqexcel["Matriz Local"]
        for i in range(8):
            for j in range(8):
                Pastatraba.cell(row=aux + 6, column=1, value='Barra : ' + str(bar + 1))
                Pastatraba.cell(row=i + aux + 7, column=j + 2, value=matriz[i, j])

    if tipo == 2:
        # Grava as matrizes de força equivalente  no arquivo excel
        Pastatraba = arqexcel["Força EQ"]
        for i in range(8):
            Pastatraba.cell(row=aux + 6, column=1, value='Barra : ' + str(bar + 1))
            Pastatraba.cell(row=i + aux + 7, column=2, value=matriz[i, 0])

    if tipo == 3:
        # Grava a matriz global com condição de contorno
        Pastatraba = arqexcel["Mat K"]
        for i in range(int(4 * nno)):
            for j in range(int(4 * nno)):
                Pastatraba.cell(row=i + 6, column=j + 1, value=matriz[i, j])

    if tipo == 4:
        # Grava Vetor de foça nodal ( Feq + Fo ) com condição de contorno
        Pastatraba = arqexcel["Mat F"]
        for i in range(1, 4 * nno + 1, 1):
            Pastatraba.cell(row=i + 6, column=1, value=int(ceil(i / 4)))
            Pastatraba.cell(row=i + 6, column=2, value=Descod[mod(i - 1, 4)])
            Pastatraba.cell(row=i + 6, column=3, value=matriz[i - 1, 0])

    # Garva a matriz global
    if tipo == 5:
        Pastatraba = arqexcel["Matriz Global"]
        for i in range(int(4 * nno)):
            for j in range(int(4 * nno)):
                Pastatraba.cell(row=i + 6, column=j + 1, value=matriz[i, j])

    if tipo == 6:
        # Grava Vetor de foça nodal ( Feq + Fo )
        Pastatraba = arqexcel["Força F"]
        for i in range(1, 4 * nno + 1, 1):
            Pastatraba.cell(row=i + 6, column=1, value=int(ceil(i / 4)))
            Pastatraba.cell(row=i + 6, column=2, value=Descod[mod(i - 1, 4)])
            Pastatraba.cell(row=i + 6, column=3, value=matriz[i - 1, 0])

    if tipo == 7:
        # Grava Vetor Deslocmentos
        Pastatraba = arqexcel["Deslocamentos"]
        for i in range(1, 4 * nno + 1, 1):
            Pastatraba.cell(row=i + 5, column=1, value=int(ceil(i / 4)))
            Pastatraba.cell(row=i + 5, column=2, value=Descod[mod(i - 1, 4)])
            Pastatraba.cell(row=i + 5, column=3, value=matriz[i - 1, 0])

    arqexcel.save(Nome_aquivo)

# -.-.-.-.-.-.-.-.-.-.-.-.-.-.-.-.-.-.-.-.-.-.-.-.-.-.-.-.-.-.-.-.-.-.-.-.-.-.-.-.-.
# Limpa o banco de dados para receber novos valores
def limpa(Nome_aquivo):
    arqexcel = load_workbook(Nome_aquivo)

    pastatraba = arqexcel['Matriz Local']
    for i in range(6, 300, 1):
        for j in range(1, 10, 1):
            pastatraba.cell(row=i, column=j, value='')

    pastatraba = arqexcel['Força EQ']
    for i in range(6, 300, 1):
        for j in range(1, 3, 1):
            pastatraba.cell(row=i, column=j, value='')

    pastatraba = arqexcel['Matriz Global']
    for i in range(6, 300, 1):
        for j in range(1, 300, 1):
            pastatraba.cell(row=i, column=j, value='')

    pastatraba = arqexcel['Força F']
    for i in range(6, 300, 1):
        for j in range(1, 4, 1):
            pastatraba.cell(row=i, column=j, value='')

    pastatraba = arqexcel['Mat K']
    for i in range(6, 300, 1):
        for j in range(1, 300, 1):
            pastatraba.cell(row=i, column=j, value='')

    pastatraba = arqexcel['Mat F']
    for i in range(6, 300, 1):
        for j in range(1, 4, 1):
            pastatraba.cell(row=i, column=j, value='')

    pastatraba = arqexcel["Deslocamentos"]
    for i in range(6, 300, 1):
        for j in range(1, 4, 1):
            pastatraba.cell(row=i, column=j, value='')

    pastatraba = arqexcel['Esforços']
    for i in range(6, 300, 1):
        for j in range(1, 10, 1):
            pastatraba.cell(row=i, column=j, value='')

    arqexcel.save(Nome_aquivo)

# -.-.-.-.-.-.-.-.-.-.-.-.-.-.-.-.-.-.-.-.-.-.-.-.-.-.-.-.-.-.-.-.-.-.-.-.-.-.-.-.-.
# Determina o numero de nós
def numero_nos(Nome_aquivo):
    arqexcel = load_workbook(Nome_aquivo)
    Pastatraba = arqexcel["Entrada Nós"]
    a = 0
    nno = 0
    i = 1
    while a == 0:
        if str(Pastatraba.cell(row=(5 + i), column=1).value) != "None":
            nno += 1
            i += 1
        else:
            a = 1

    return nno

# -.-.-.-.-.-.-.-.-.-.-.-.-.-.-.-.-.-.-.-.-.-.-.-.-.-.-.-.-.-.-.-.-.-.-.-.-.-.-.-.-.
# Determina o numero de Barras
def numero_barras(Nome_aquivo):
    arqexcel = load_workbook(Nome_aquivo)
    Pastatraba = arqexcel["Entrada Barras"]
    a = 0
    nbar = 0
    i = 1
    while a == 0:
        if str(Pastatraba.cell(row=5 + i, column=1).value) != "None":
            nbar += 1
            i += 1
        else:
            a = 1
    return nbar

# -.-.-.-.-.-.-.-.-.-.-.-.-.-.-.-.-.-.-.-.-.-.-.-.-.-.-.-.-.-.-.-.-.-.-.-.-.-.-.-.-.
# Determina o comprimento das Barras
def comprimento(xi, xf, yi, yf):
    compr = sqrt((xf - xi) ** 2 + (yf - yi) ** 2)
    return compr

# -.-.-.-.-.-.-.-.-.-.-.-.-.-.-.-.-.-.-.-.-.-.-.-.-.-.-.-.-.-.-.-.-.-.-.-.-.-.-.-.-.
# Determina o angulo das Barras
def angulo(xi, xf, yi, yf):
    compr = sqrt((xf - xi) ** 2 + (yf - yi) ** 2)
    vcos = (xf - xi) / compr
    vsen = (yf - yi) / compr
    return vsen, vcos

# -.-.-.-.-.-.-.-.-.-.-.-.-.-.-.-.-.-.-.-.-.-.-.-.-.-.-.-.-.-.-.-.-.-.-.-.-.-.-.-.-.
# Monta a Matriz cinamática
def matriz_cinematica(dsen, dcos):
    mcin = np.zeros((8, 8))

    mcin[0, 0] = 1
    mcin[4, 4] = 1

    mcin[1, 1] = dcos
    mcin[1, 2] = dsen
    mcin[2, 1] = -dsen
    mcin[2, 2] = dcos
    mcin[3, 3] = 1

    mcin[5, 5] = dcos
    mcin[5, 6] = dsen
    mcin[6, 5] = -dsen
    mcin[6, 6] = dcos
    mcin[7, 7] = 1

    return mcin

# -.-.-.-.-.-.-.-.-.-.-.-.-.-.-.-.-.-.-.-.-.-.-.-.-.-.-.-.-.-.-.-.-.-.-.-.-.-.-.-.-.
# Monta a Matriz de Rigiez Local
def Smge(tipo, G, It, E, I, A, L, kf,P):
    mge = np.zeros((8, 8))

    ## Matriz de rigidez local da barra com 8 gruas de liberdade
    # barra bi engastada
    if tipo == 11:
        mge[0, 0] = G * It / L
        mge[0, 4] = -G * It / L
        mge[1, 1] = E * A / L
        mge[1, 5] = -E * A / L
        mge[2, 2] = 12 * E * I / L ** 3
        mge[2, 3] = 6 * E * I / L ** 2
        mge[2, 6] = -12 * E * I / L ** 3
        mge[2, 7] = 6 * E * I / L ** 2
        mge[3, 2] = 6 * E * I / L ** 2
        mge[3, 3] = 4 * E * I / L
        mge[3, 6] = -6 * E * I / L ** 2
        mge[3, 7] = 2 * E * I / L
        mge[4, 0] = -G * It / L
        mge[4, 4] = G * It / L
        mge[5, 1] = -E * A / L
        mge[5, 5] = E * A / L
        mge[6, 2] = -12 * E * I / L ** 3
        mge[6, 3] = -6 * E * I / L ** 2
        mge[6, 6] = 12 * E * I / L ** 3
        mge[6, 7] = -6 * E * I / L ** 2
        mge[7, 2] = 6 * E * I / L ** 2
        mge[7, 3] = 2 * E * I / L
        mge[7, 6] = -6 * E * I / L ** 2
        mge[7, 7] = 4 * E * I / L

    # barra engastada rotulada
    if tipo == 10:
        mge[0, 0] = G * It / L
        mge[0, 4] = -G * It / L
        mge[1, 1] = E * A / L
        mge[1, 5] = -E * A / L
        mge[2, 2] = 3 * E * I / L ** 3
        mge[2, 3] = 3 * E * I / L ** 2
        mge[2, 6] = -3 * E * I / L ** 3

        mge[3, 2] = 3 * E * I / L ** 2
        mge[3, 3] = 3 * E * I / L
        mge[3, 6] = -3 * E * I / L ** 2

        mge[4, 0] = -G * It / L
        mge[4, 4] = G * It / L
        mge[5, 1] = -E * A / L
        mge[5, 5] = E * A / L
        mge[6, 2] = -3 * E * I / L ** 3
        mge[6, 3] = -3 * E * I / L ** 2
        mge[6, 6] = 3 * E * I / L ** 3

    # Barra rotulada engastada
    if tipo == 1:
        mge[0, 0] = G * It / L
        mge[0, 4] = -G * It / L
        mge[1, 1] = E * A / L
        mge[1, 5] = -E * A / L
        mge[2, 2] = 3 * E * I / L ** 3

        mge[2, 6] = -3 * E * I / L ** 3
        mge[2, 7] = 3 * E * I / L ** 2

        mge[4, 0] = -G * It / L
        mge[4, 4] = G * It / L
        mge[5, 1] = -E * A / L
        mge[5, 5] = E * A / L
        mge[6, 2] = -3 * E * I / L ** 3

        mge[6, 6] = 3 * E * I / L ** 3
        mge[6, 7] = -3 * E * I / L ** 2
        mge[7, 2] = 3 * E * I / L ** 2

        mge[7, 6] = -3 * E * I / L ** 2
        mge[7, 7] = 3 * E * I / L

    # Barra bi articulada
    if tipo == 0:
        mge[1, 1] = E * A / L
        mge[1, 5] = -E * A / L

        mge[5, 1] = -E * A / L
        mge[5, 5] = E * A / L

        # Contrubuição das molas distrubuidas
    mge[2, 2] = mge[2, 2] + (kf * L / 420) * 156
    mge[2, 3] = mge[2, 3] + (kf * L / 420) * 22 * L
    mge[2, 6] = mge[2, 6] + (kf * L / 420) * 54
    mge[2, 7] = mge[2, 7] + (kf * L / 420) * (-13 * L)
    mge[3, 2] = mge[3, 2] + (kf * L / 420) * 22 * L
    mge[3, 3] = mge[3, 3] + (kf * L / 420) * 4 * L ** 2
    mge[3, 6] = mge[3, 6] + (kf * L / 420) * 13 * L
    mge[3, 7] = mge[3, 7] + (kf * L / 420) * (-3 * L ** 2)
    mge[6, 2] = mge[6, 2] + (kf * L / 420) * 54
    mge[6, 3] = mge[6, 3] + (kf * L / 420) * 13 * L
    mge[6, 6] = mge[6, 6] + (kf * L / 420) * 156
    mge[6, 7] = mge[6, 7] + (kf * L / 420) * (-22 * L)
    mge[7, 2] = mge[7, 2] + (kf * L / 420) * (-13 * L)
    mge[7, 3] = mge[7, 3] + (kf * L / 420) * (- 3 * L ** 2)
    mge[7, 6] = mge[7, 6] + (kf * L / 420) * (-22 * L)
    mge[7, 7] = mge[7, 7] + (kf * L / 420) * 4 * L ** 2

    # contribuição de carga concentrada
    mge[2, 2] = mge[2, 2] + (P / (30 * L)) * 36
    mge[2, 3] = mge[2, 3] + (P / (30 * L)) * 3 * L
    mge[2, 6] = mge[2, 6] + (P / (30 * L)) * (-36)
    mge[2, 7] = mge[2, 7] + (P / (30 * L)) * 3 * L
    mge[3, 2] = mge[3, 2] + (P / (30 * L)) * 3 * L
    mge[3, 3] = mge[3, 3] + (P / (30 * L)) * 4 * L ** 2
    mge[3, 6] = mge[3, 6] + (P / (30 * L)) * (-3 * L)
    mge[3, 7] = mge[3, 7] + (P / (30 * L)) * (-L ** 2)
    mge[6, 2] = mge[6, 2] + (P / (30 * L)) * (-36)
    mge[6, 3] = mge[6, 3] + (P / (30 * L)) * (-3 * L)
    mge[6, 6] = mge[6, 6] + (P / (30 * L)) * 36
    mge[6, 7] = mge[6, 7] + (P / (30 * L)) * (-3 * L)
    mge[7, 2] = mge[7, 2] + (P / (30 * L)) * 3 * L
    mge[7, 3] = mge[7, 3] + (P / (30 * L)) * (-L ** 2)
    mge[7, 6] = mge[7, 6] + (P / (30 * L)) * (-3 * L)
    mge[7, 7] = mge[7, 7] + (P / (30 * L)) * 4 * L ** 2

    return mge

# -.-.-.-.-.-.-.-.-.-.-.-.-.-.-.-.-.-.-.-.-.-.-.-.-.-.-.-.-.-.-.-.-.-.-.-.-.-.-.-.-.
# Monta o Vetor de Força Nodal
def forcaequivale(tipo, compr, qix, qfx, qiy, qfy, qig, qfg):
    po = np.zeros((8, 1))

    # carga nodal
    # barra bi engastada
    if tipo == 11:
        po[0, 0] = (compr / 3.0) * (qig + 0.5 * qfg)
        po[1, 0] = (compr / 3.0) * (qix + 0.5 * qfx)
        po[2, 0] = (compr / 60.0) * (21 * qiy + 9 * qfy)
        po[3, 0] = (compr / 60.0) * (3 * qiy * compr + 2 * qfy * compr)
        po[4, 0] = (compr / 3.0) * (0.5 * qig + qfg)
        po[5, 0] = (compr / 3.0) * (0.5 * qix + qfx)
        po[6, 0] = (compr / 60.0) * (9 * qiy + 21 * qfy)
        po[7, 0] = (-compr / 60.0) * (2 * qiy * compr + 3 * qfy * compr)

    # barra Engastada apoioada
    if tipo == 10:
        po[0, 0] = (compr / 3.0) * (qig + 0.5 * qfg)
        po[1, 0] = (compr / 3.0) * (qix + 0.5 * qfx)
        po[2, 0] = (compr / 15.0) * (6 * qiy + 3.375 * qfy)
        po[3, 0] = (compr / 15.0) * (qiy * compr + 0.875 * qfy * compr)
        po[4, 0] = (compr / 3.0) * (0.5 * qig + qfg)
        po[5, 0] = (compr / 3.0) * (0.5 * qix + qfx)
        po[6, 0] = (compr / 15.0) * (1.5 * qiy + 4.125 * qfy)
        po[7, 0] = 0

    # barra apoiada Engastada
    if tipo == 1:
        po[0, 0] = (compr / 3.0) * (qig + 0.5 * qfg)
        po[1, 0] = (compr / 3.0) * (qix + 0.5 * qfx)
        po[2, 0] = (compr / 15.0) * (4.125 * qiy + 1.5 * qfy)
        po[3, 0] = 0
        po[4, 0] = (compr / 3.0) * (0.5 * qig + qfg)
        po[5, 0] = (compr / 3.0) * (0.5 * qix + qfx)
        po[6, 0] = (compr / 15.0) * (3.375 * qiy + 6 * qfy)
        po[7, 0] = (-compr / 15.0) * (0.875 * qiy * compr + qfy * compr)

    return po

# -.-.-.-.-.-.-.-.-.-.-.-.-.-.-.-.-.-.-.-.-.-.-.-.-.-.-.-.-.-.-.-.-.-.-.-.-.-.-.-.-.
# Monta o Vetor de Força Nodal
def forcanodal(Mfor, Fo, noi, nof, po):
    Fnod = -po + Fo
    #  Carga no nó inical
    Mfor[4 * noi - 4, 0] = Mfor[4 * noi - 4, 0] + Fnod[0, 0]
    Mfor[4 * noi - 3, 0] = Mfor[4 * noi - 3, 0] + Fnod[1, 0]
    Mfor[4 * noi - 2, 0] = Mfor[4 * noi - 2, 0] + Fnod[2, 0]
    Mfor[4 * noi - 1, 0] = Mfor[4 * noi - 1, 0] + Fnod[3, 0]
    # Carga no nó final
    Mfor[4 * nof - 4, 0] = Mfor[4 * nof - 4, 0] + Fnod[4, 0]
    Mfor[4 * nof - 3, 0] = Mfor[4 * nof - 3, 0] + Fnod[5, 0]
    Mfor[4 * nof - 2, 0] = Mfor[4 * nof - 2, 0] + Fnod[6, 0]
    Mfor[4 * nof - 1, 0] = Mfor[4 * nof - 1, 0] + Fnod[7, 0]

    return Mfor

# -.-.-.-.-.-.-.-.-.-.-.-.-.-.-.-.-.-.-.-.-.-.-.-.-.-.-.-.-.-.-.-.-.-.-.-.-.-.-.-.-.
# Impoe a vinculação
def vinculacao(matrizk, matrizf, no, rx, ry, rz, rg):
    n = len(matrizk)  # numero de linhas da matriz

    for ii in range(1, int(n / 4) + 1, 1):

        if ii == no:

            if rg == 1:
                matrizf[4 * ii - 4, 0] = 0

                for jj in range(1, n + 1, 1):
                    if 4 * ii - 3 == jj:
                        matrizk[4 * ii - 4, jj - 1] = 1
                    else:
                        matrizk[4 * ii - 4, jj - 1] = 0
                        matrizk[jj - 1, 4 * ii - 4] = 0

            if rx == 1:
                matrizf[4 * ii - 3, 0] = 0

                for jj in range(1, n + 1, 1):
                    if 4 * ii - 2 == jj:
                        matrizk[4 * ii - 3, jj - 1] = 1
                    else:
                        matrizk[4 * ii - 3, jj - 1] = 0
                        matrizk[jj - 1, 4 * ii - 3] = 0
            if ry == 1:
                matrizf[4 * ii - 2, 0] = 0

                for jj in range(1, n + 1, 1):
                    if 4 * ii - 1 == jj:
                        matrizk[4 * ii - 2, jj - 1] = 1
                    else:
                        matrizk[4 * ii - 2, jj - 1] = 0
                        matrizk[jj - 1, 4 * ii - 2] = 0

            if rz == 1:
                matrizf[4 * ii - 1, 0] = 0

                for jj in range(1, n + 1, 1):
                    if 4 * ii == jj:
                        matrizk[4 * ii - 1, jj - 1] = 1
                    else:
                        matrizk[4 * ii - 1, jj - 1] = 0
                        matrizk[jj - 1, 4 * ii - 1] = 0

    return matrizk, matrizf

# -.-.-.-.-.-.-.-.-.-.-.-.-.-.-.-.-.-.-.-.-.-.-.-.-.-.-.-.-.-.-.-.-.-.-.-.-.-.-.-.-.
# Monta-se a Matriz Global
def matrizglobal(matrizk, mge, noi, nof):
    for k in [1, 2, 3, 4]:
        for i in [1, 2, 3, 4]:
            for j in [1, 2, 3, 4]:
                matrizk[int(4 * ((2 - ceil(k / 2)) * noi + (ceil(k / 2) - 1) * nof) - 5 + i),
                        4 * (mod(k, 2) * noi + mod(k - 1, 2) * nof) - 5 + j] = \
                    matrizk[int(4 * ((2 - ceil(k / 2)) * noi + (ceil(k / 2) - 1) * nof) - 5 + i),
                            4 * ((mod(k, 2)) * noi + (mod(k - 1, 2)) * nof) - 5 + j] + \
                    mge[int(4 * (ceil(k / 2)) - 5 + i), 4 * (mod(k - 1, 2) + 1) - 5 + j]
    return matrizk


# -.-.-.-.-.-.-.-.-.-.-.-.-.-.-.-.-.-.-.-.-.-.-.-.-.-.-.-.-.-.-.-.-.-.-.-.-.-.-.-.-.
# Calculo dos Esforços solicitantes
def Esforcos_solicitantes(Nome_aquivo):
    global Matno, nbar, Matbar, Carbar, desl,delta
    # Calcula os esforços solicitantes
    for bar in range(int(nbar)):
        # Tipo de Calculo
        tipo = int(Matbar[bar, 1])

        # nó inicial e final da barra
        noi = int(Matbar[bar, 2])
        nof = int(Matbar[bar, 3])

        # Coordenadas iniciais e finais da barra
        xi = Matno[noi - 1, 1]
        yi = Matno[noi - 1, 2]
        xf = Matno[nof - 1, 1]
        yf = Matno[nof - 1, 2]

        # Propriedades das Barras
        A = Matbar[bar, 4]  # área da barra
        I = Matbar[bar, 5]  # inércia flexional
        It = Matbar[bar, 6]  # inércia torcional
        E = Matbar[bar, 7]  # Módulo de Elasticidade
        poisson = Matbar[bar, 8]  # poisson
        ptor = Matbar[bar, 10]  # porcentagem da rigidez a torção
        G = E / (2 * (1 + poisson))

        # Carregamentos na Barra
        qix = Carbar[bar, 1]  # carga inicial paralelo a barra
        qfx = Carbar[bar, 2]  # carga final paralelo a barra
        qiy = Carbar[bar, 3]  # carga inicial perpendicular a barra
        qfy = Carbar[bar, 4]  # carga final perpendicular a barra
        qig = Carbar[bar, 5]  # carga de torcão inicial na barra
        qfg = Carbar[bar, 6]  # carga de torçao final na barra

        # montagem da matriz de rigidez e vetor de carga nodal
        L = comprimento(xi, xf, yi, yf)  # comprimento da Barra
        vsen, vcos = angulo(xi, xf, yi, yf)  # Angulo da Barra
        mge = Smge(tipo, G, It * ptor / 100, E, I, A, L, 0, 0)  # Matriz de Rigidez local


        # Deslocamentos na barra rotacionado para o sistema global
        delta[0, 0] = desl[4 * noi - 4, 0]
        delta[1, 0] = desl[4 * noi - 3, 0] * vcos + desl[4 * noi - 2, 0] * vsen
        delta[2, 0] = -desl[4 * noi - 3, 0] * vsen + desl[4 * noi - 2, 0] * vcos
        delta[3, 0] = desl[4 * noi - 1, 0]
        delta[4, 0] = desl[4 * nof - 4, 0]
        delta[5, 0] = desl[4 * nof - 3, 0] * vcos + desl[4 * nof - 2, 0] * vsen
        delta[6, 0] = -desl[4 * nof - 3, 0] * vsen + desl[4 * nof - 2, 0] * vcos
        delta[7, 0] = desl[4 * nof - 1, 0]

        # Matriz de força equivalente e Matriz de Força nodal
        po = forcaequivale(tipo, L, qix, qfx, qiy, qfy, qig, qfg)

        esf = po + mge.dot(delta)  # esforços na barra

        # ajustes de sinais
        esf[1, 0] = -esf[1, 0]
        esf[6, 0] = -esf[6, 0]
        esf[7, 0] = -esf[7, 0]

        # esforços solicitantes
        arqexcel = load_workbook(Nome_aquivo)
        Pastatraba = arqexcel["Esforços"]

        Pastatraba.cell(row=bar + 6, column=1, value='Barra : ' + str(bar))
        for i in range(8):
            Pastatraba.cell(row=bar + 6, column=1, value='Barra : ' + str(bar + 1))
            Pastatraba.cell(row=bar + 6, column=i + 2, value=esf[i, 0])

        arqexcel.save(Nome_aquivo)
    messagebox.showinfo('ABOUT', 'Cálculos finalizados com sucesso')


