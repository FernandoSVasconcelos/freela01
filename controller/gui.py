import numpy as np
import subprocess
import sys

from numpy import linalg 
from openpyxl import * 
from tkinter import *
from tkinter import messagebox
from tkinter import scrolledtext
from tkinter import filedialog
from tkinter import ttk
from datetime import datetime
from math import *

sys.path.insert(0, "/home/ubuntu/Downloads/freela01/controller")
from functions import *

Nome_aquivo = ''
aberto = 0
Titulo = 'TITULO DO PROJETO'

class Tela_About:
    def __init__(self, Raiz, Original):
        self.TelAbout = Raiz
        self.Teloriginal = Original

        self.TelAbout.title('Sobre')
        self.TelAbout.geometry("400x250+400+150")
        self.TelAbout.resizable(0, 0)

        self.canvas = Canvas(self.TelAbout, width=380, height=230, bg='#EEE8AA')
        self.canvas.place(x=10, y=10)

        self.lb1 = Label(self.TelAbout, text="PUC Minas - Poços de Caldas", font=('Times', '13', 'bold'), bg='#EEE8AA')
        self.lb1.place(x=80, y=20)

        self.lb1 = Label(self.TelAbout, text="ESCREVER SOBRE O METODO UTILIZADO", font=('Times', '11'),bg='#EEE8AA')
        self.lb1.place(x=50, y=80)

        self.lb1 = Label(self.TelAbout, text="Data : 01/02/2022", font=('Times', '11'), bg='#EEE8AA')
        self.lb1.place(x=50, y=150)

        self.lb1 = Label(self.TelAbout, text="Versão: V1.0", font=('Times', '11'), bg='#EEE8AA')
        self.lb1.place(x=50, y=130)

        self.lb1 = Label(self.TelAbout, text="Programadores:  -Karolina Ribeiro\n\t       -Lucas Amaro", font=('Times', '11'), bg='#EEE8AA')
        self.lb1.place(x=50, y=200)

class GuiShow:
    def __init__(self, Raiz, Original, valores_classe):
        self.gui_show = Raiz
        self.Teloriginal = Original

        self.gui_show.title('Mostrar Valores')
        self.gui_show.geometry("400x250+400+150")
        self.gui_show.resizable(0, 0)

        self.canvas = Canvas(self.gui_show, width=400, height=250, bg='#EEE8AA')
        self.canvas.place(x=0, y=0)

        self.lb1 = Label(self.gui_show, text = f"Formato do canal: {valores_classe.Var1.get()}", font=('Times', '11'), bg='#EEE8AA')
        self.lb1.place(x=50, y=30)

        self.lb1 = Label(self.gui_show, text = f"Vazão: {valores_classe.vazão}", font=('Times', '11'),bg='#EEE8AA')
        self.lb1.place(x=50, y=50)

        self.lb1 = Label(self.gui_show, text = f"Rugosidade: {valores_classe.rugosidade}", font=('Times', '11'), bg='#EEE8AA')
        self.lb1.place(x=50, y=70)

        self.lb1 = Label(self.gui_show, text = f"Teta: {valores_classe.teta}", font=('Times', '11'), bg='#EEE8AA')
        self.lb1.place(x=50, y=90)

        self.lb1 = Label(self.gui_show, text = f"d: {valores_classe.d}", font=('Times', '11'), bg='#EEE8AA')
        self.lb1.place(x=50, y=110)

        self.lb1 = Label(self.gui_show, text = f"Base menor: {valores_classe.base_menor}", font=('Times', '11'), bg='#EEE8AA')
        self.lb1.place(x=50, y=130)

        self.lb1 = Label(self.gui_show, text = f"z: {valores_classe.z}", font=('Times', '11'), bg='#EEE8AA')
        self.lb1.place(x=50, y=150)

class Tela_Principal:
    def __init__(self, Raiz):
        self.vazão = 0
        self.rugosidade = 0
        self.teta = 0
        self.d = 0
        self.base_menor = 0
        self.z = 0
        self.path = ''

        self.TelPrin = Raiz
        self.agora = datetime.now()
        
        self.TelPrin.title(Titulo)
        self.TelPrin.geometry("1000x545+150+10") 
        self.TelPrin.resizable(0, 0)

        self.canvas = Canvas(self.TelPrin, width=200, height=30, bg='#DCDCDC')
        self.canvas.place(x=0, y=510)

        self.canvas = Canvas(self.TelPrin, width=800, height=30, bg='Silver')
        self.canvas.place(x=200, y=510)

        self.Relogio = Label(self.TelPrin, text=self.agora.strftime('%d/%m/%Y       %H:%M:%S'),
                                           font=('Times', '10', 'bold'),
                                           bg='#DCDCDC',
                                           relief = "sunken")
        self.Relogio.place(x=5, y=515)

        self.ajud = Label(self.TelPrin, text="ESCREVER SOBRE O OBJETIVO DO PROJETO - Versão : V1.0                          " ,
                                        font=('Times', '10', 'bold'),
                                        bg='Silver',
                                        relief = "sunken")
        self.ajud.place(x=210, y=517)


        # ////Cria meunu superior  -  Principal
        self.Menusup = Menu(self.TelPrin, tearoff=0)  # Cria um menu superior
        self.TelPrin.configure(menu=self.Menusup)  # Configura a tela inicial para o menu

        # ////Cria meunu superior  -  Arquivo
        self.arquivo = Menu(self.Menusup, tearoff=0)  # Cria um menu dentro do menu superior Arquivo
        self.arquivo.add_command(label="Abrir", command=self.Sub_Abrir)  # Adiciona Abrir ao menu
        self.arquivo.add_command(label="Sair", command=self.Sub_Sair)  # Adiciona sair no sub menu
        self.Menusup.add_cascade(label="Arquivo", menu=self.arquivo)  # Adiciona  cascata

        # //// Cria  Menu superior - Saída de dados
        self.Menusup.add_command(label="Resultados", command=self.Sub_Resultados)

        # //// Cria  Menu superior - About
        self.Menusup.add_command(label="Sobre", command=self.Sub_About)

        self.Menusup.add_command(label="Mostrar Valores", command=self.show_values)

        # //// Cria  Menu superior - Sair
        self.Menusup.add_command(label="Sair", command=self.Sub_Sair)

        self.Var1 = StringVar(value="1")
        
        ChkBttn = Radiobutton(self.TelPrin, width = 15, variable = self.Var1, text = 'Retangular', value = 'Retangular')
        ChkBttn.place(x = 12, y = 25)
        
        ChkBttn2 = Radiobutton(self.TelPrin, width = 15, variable = self.Var1, text = 'Triangular', value = 'Triangular')
        ChkBttn2.place(x = 10, y = 55)

        ChkBttn3 = Radiobutton(self.TelPrin, width = 15, variable = self.Var1, text = 'Trapezoidal Simétrico', value = 'Trapezoidal Simétrico')
        ChkBttn3.place(x = 40, y = 85)

        button1 = Button(self.TelPrin, text = "Calcular", command=self.calcular)
        button1.place(x = 250, y = 60)

        #  atualização de tela
        self.alteracao()

        self.TelPrin.protocol("WM_DELETE_WINDOW", self.on_closing_Principal)

    # ************************************************Sub rotinas *******************************************
    # Atualização do Relógio
    def alteracao(self):
        self.agora = datetime.now()
        self.Relogio['text'] = self.agora.strftime('%d/%m/%Y       %H:%M:%S')
        self.TelPrin.after(1000, self.alteracao)

    def Sub_About(self):
        self.Telabout = Toplevel(self.TelPrin)
        Tela_About(self.Telabout, self.TelPrin)
        self.Telabout.protocol("WM_DELETE_WINDOW", self.on_close_about)

    def show_values(self):
        self.close_show = Toplevel(self.TelPrin)
        GuiShow(self.close_show, self.TelPrin, self)
        self.close_show.protocol("WM_DELETE_WINDOW", self.on_close_show)

    def on_close_show(self):
        self.close_show.destroy()

    def on_close_about(self):
        self.Telabout.destroy()

    def on_closing_Principal(self):
     self.Sub_Sair()

    def Sub_Sair(self):
        self.escolha = messagebox.askquestion('CONFIRMAÇÃO!!!', 'Tem certeza que deseja Sair?')
        if self.escolha == 'yes':
            self.TelPrin.destroy()

    def Sub_Abrir(self):
        global  Nome_aquivo , Titulo

        entrada_arquivo = filedialog.askopenfile(initialdir="/home/ubuntu/Documentos/freela01/files", title="Abrir arquivo",
                                               filetypes=(("Entrada", "*.XLSX"), ("all files", "*.*")))
        if entrada_arquivo != None:
            Nome_aquivo = entrada_arquivo.name
            Titulo = Nome_aquivo
            excel_file = load_workbook(Nome_aquivo)
            self.path = excel_file["Entrada de Dados"]

            self.vazão = self.path.cell(row = 10, column = 1).value
            self.rugosidade = self.path.cell(row = 13, column = 1).value
            self.teta = self.path.cell(row = 7, column = 3).value
            self.d = self.path.cell(row = 10, column = 3).value
            self.base_menor = self.path.cell(row = 13, column = 3).value
            self.z = self.path.cell(row = 7, column = 5).value

    def calcular(self):
      global Nome_aquivo , aberto
      if Nome_aquivo == '' :
          messagebox.showinfo('ABOUT', 'Abra um arquivo Primeiro')
      else:
        get_valores(self)

    def Sub_Resultados(self):
        global Nome_aquivo
        if Nome_aquivo == '':
            messagebox.showinfo('ABOUT', 'Abra um arquivo Primeiro...\n')
        else:
            t = subprocess.Popen(Nome_aquivo, shell=True)
            t.wait()

if __name__ == '__main__':
    Principal = Tk()
    Tela_Principal(Principal)
    Principal.mainloop()